import logging
from typing import Any

import openpyxl

from app.rag.parsers.base import BaseParser

logger = logging.getLogger("zam-ai-core-api.xlsx-parser")


class XlsxParser(BaseParser):
    def parse(self, file_path: str) -> list[dict[str, Any]]:
        logger.info(f"Parsing XLSX file: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sections = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if not header_row:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            headers_lower = [h.lower().strip() for h in headers]

            rows_data = [row for row in rows_iter]

            if self._is_atc_classification(headers_lower):
                sections.extend(self._parse_atc_classification(rows_data, headers, sheet_name))
            else:
                sections.extend(self._parse_generic_sheet(rows_data, headers, sheet_name))

        wb.close()
        return sections

    def _is_atc_classification(self, headers: list[str]) -> bool:
        return "generic name" in headers and "generic atc code" in headers

    def _parse_atc_classification(
        self,
        rows: list[tuple],
        headers: list[str],
        sheet_name: str,
    ) -> list[dict[str, Any]]:
        sections = []
        seen_levels: dict[str, set[str]] = {}

        for row in rows:
            if not row or all(v is None for v in row):
                continue
            row_dict = dict(zip(headers, [str(v or "").strip() for v in row], strict=False))
            generic_name = row_dict.get("Generic Name", "").strip()
            atc_code = row_dict.get("Generic ATC Code", "").strip()
            level4_name = row_dict.get("Level4 Name", "").strip()
            level4_code = row_dict.get("Level4 Code", "").strip()
            level3_name = row_dict.get("Level3 Name", "").strip()
            level2_name = row_dict.get("Level2 Name", "").strip()
            level1_name = row_dict.get("Level1 Name", "").strip()
            level1_alias = row_dict.get("Level1 Alias", "").strip()

            if level1_name:
                if level1_name not in seen_levels:
                    seen_levels[level1_name] = set()
                    sections.append({
                        "text_content": (
                            f"ATC Level 1: {level1_name}\n"
                            f"Code: {row_dict.get('Level1 Code', '')}\n"
                            f"Alias: {level1_alias}"
                        ),
                        "section_path": f"ATC / {level1_name}",
                        "page_number": None,
                        "metadata": {"chunk_type": "atc_level1"},
                    })
            if level2_name and level2_name not in seen_levels.get(level1_name, set()):
                if level1_name not in seen_levels:
                    seen_levels[level1_name] = set()
                seen_levels[level1_name].add(level2_name)
                sections.append({
                    "text_content": (
                        f"ATC Level 2: {level2_name}\n"
                        f"Code: {row_dict.get('Level2 Code', '')}\n"
                        f"Parent: {level1_name}"
                    ),
                    "section_path": f"ATC / {level1_name} / {level2_name}",
                    "page_number": None,
                    "metadata": {"chunk_type": "atc_level2"},
                })

            if not generic_name or not atc_code:
                continue

            lines = [
                f"Generic Name: {generic_name}",
                f"ATC Code: {atc_code}",
            ]
            if level1_name:
                lines.append(f"Anatomical Group: {level1_name} ({row_dict.get('Level1 Code', '')})")
            if level2_name:
                lines.append(f"Therapeutic Group: {level2_name} ({row_dict.get('Level2 Code', '')})")
            if level3_name:
                lines.append(f"Pharmacological Group: {level3_name} ({row_dict.get('Level3 Code', '')})")
            if level4_name:
                lines.append(f"Chemical Group: {level4_name} ({level4_code})")

            sections.append({
                "text_content": "\n".join(lines),
                "section_path": f"ATC / {level1_name or 'Unknown'} / {level2_name or 'Unknown'} / {generic_name}",
                "page_number": None,
                "metadata": {
                    "generic_name": generic_name,
                    "atc_code": atc_code,
                    "chunk_type": "atc_entry",
                },
            })
        return sections

    def _parse_generic_sheet(
        self,
        rows: list[tuple],
        headers: list[str],
        sheet_name: str,
    ) -> list[dict[str, Any]]:
        sections = []
        for index, row in enumerate(rows):
            if not row or all(v is None for v in row):
                continue
            row_dict = dict(zip(headers, [str(v or "").strip() for v in row], strict=False))
            text_parts = [f"{k}: {v}" for k, v in row_dict.items() if v]
            if not text_parts:
                continue
            sections.append({
                "text_content": "\n".join(text_parts),
                "section_path": f"{sheet_name} / Row {index}",
                "page_number": None,
                "metadata": {"type": "xlsx_row", "sheet": sheet_name},
            })
        return sections
